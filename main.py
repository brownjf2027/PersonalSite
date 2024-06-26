import datetime
import json
import logging
from os import environ
from datetime import datetime
import pytz
import openpyxl

import todos
from ncaa import get_games
import re

from flask import Flask, request, render_template, redirect, url_for, flash
from flask_bootstrap import Bootstrap5
from flask_login import UserMixin, login_user, LoginManager, login_required, current_user, logout_user
from flask_sqlalchemy import SQLAlchemy
from flask_wtf import FlaskForm
from sqlalchemy import Integer, String, Date, LargeBinary
from sqlalchemy.orm import Mapped, mapped_column
from werkzeug.security import check_password_hash
from wtforms import StringField, SubmitField, TextAreaField, DateField, HiddenField, FileField
from wtforms.validators import DataRequired
from flask_ckeditor import CKEditor, CKEditorField


SALT_ROUNDS = 16
TODO_FILE = 'static/files/todo.json'
# class Base(DeclarativeBase):
#     pass


app = Flask(__name__)
app.logger.setLevel(logging.DEBUG)
bootstrap = Bootstrap5(app)
app.config['SECRET_KEY'] = environ.get('SECRET_KEY')
app.config['CKEDITOR_ENABLE_CODESNIPPET'] = True
app.config['CKEDITOR_CODE_THEME'] = 'monokai_sublime'
ckeditor = CKEditor(app)

# CONNECT TO DB

# app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///posts.db'
app.config['SQLALCHEMY_DATABASE_URI'] = environ.get("DB_URI")
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False  # silence the deprecation warning
db = SQLAlchemy(app)


class User(UserMixin, db.Model):
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    email: Mapped[str] = mapped_column(String(100), unique=True)
    password: Mapped[str] = mapped_column(String(500))
    name: Mapped[str] = mapped_column(String(1000))


class Post(db.Model):
    id: Mapped[int] = mapped_column(Integer, primary_key=True)
    date: Mapped[datetime.date] = mapped_column(Date)
    title: Mapped[str] = mapped_column(String(250))
    blurb: Mapped[str] = mapped_column(String(500))
    body: Mapped[str] = mapped_column(String(2000))


class Image(db.Model):
    id = mapped_column(Integer, primary_key=True)
    name = mapped_column(String)
    data = mapped_column(LargeBinary)
    post_id = db.Column(db.Integer, db.ForeignKey('post.id'))


class Snippet(db.Model):
    id = mapped_column(Integer, primary_key=True)
    name = mapped_column(String)
    data = mapped_column(String)
    post_id = db.Column(db.Integer, db.ForeignKey('post.id'))


with app.app_context():
    db.create_all()
    posts = Post.query.all()
    for post in posts:
        print(post.title)

    users = User.query.all()

app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload size

# Configure Flask-Login's Login Manager
login_manager = LoginManager()
login_manager.init_app(app)


# Create a user_loader callback
@login_manager.user_loader
def load_user(user_id):
    return db.get_or_404(User, user_id)


class PostForm(FlaskForm):
    post_id = HiddenField('Post ID:')
    date = DateField('Date')
    title = StringField('Title:', validators=[DataRequired()])
    blurb = TextAreaField('Blurb:', validators=[DataRequired()], render_kw={"style": "width: 500px;"})
    body = CKEditorField(label="Body:", validators=[DataRequired()], render_kw={"style": "height: 250px;"})
    # body = TextAreaField('Body:', validators=[DataRequired()], render_kw={"style": "height: 250px;"})
    submit = SubmitField("Submit Post", render_kw={"class": "btn btn-primary"})


class XlForm(FlaskForm):
    excel_doc = FileField(f"Choose Excel file for today's games", validators=[DataRequired()])
    submit = SubmitField('Upload Sheet')


@app.route("/post", methods=['GET', 'POST'])
@login_required
def post():
    post_id = request.args.get('edit_post_id')
    form = PostForm()

    if form.validate_on_submit():
        if post_id:
            # If post ID is provided, it's an update operation
            post = Post.query.get_or_404(post_id)
            post.id = post.id
            post.date = form.date.data
            post.title = form.title.data
            post.blurb = form.blurb.data
            post.body = form.body.data

            db.session.commit()
            flash('Saved!')
            return redirect(url_for('show_post', post_id=post.id, logged_in=current_user.is_authenticated))

        else:
            new_post = Post(
                date=form.date.data,
                title=form.title.data,
                blurb=form.blurb.data,
                body=form.body.data
            )
            db.session.add(new_post)
            db.session.commit()

            # snip = request.data['snip']

            # Check if the 'image_file' was included in the request
            if 'image_file' in request.files:
                file = request.files['image_file']
                # Check if a file was selected and if it has an allowed extension
                if file and allowed_file(file.filename):
                    file_data = file.read()
                    img_file = Image(name=file.filename, data=file_data, post_id=new_post.id)  # Use new_post.id
                    db.session.add(img_file)
                    db.session.commit()

            flash('Saved!')
            return redirect(url_for('post', post_id=new_post.id,
                                    logged_in=current_user.is_authenticated))  # Redirect with new_post.id

    if post_id:
        # If post ID is provided, fetch the existing post data for editing
        post = Post.query.get_or_404(post_id)
        form.post_id.data = post.id
        form.date.data = post.date
        form.title.data = post.title
        form.blurb.data = post.blurb
        form.body.data = post.body

    return render_template("add_post.html", form=form, logged_in=current_user.is_authenticated)


def allowed_file(filename):
    allowed_extensions = {'jpg', 'gif', 'png'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions


@app.route("/post/<int:post_id>")
def show_post(post_id):
    requested_post = Post.query.filter_by(id=post_id).first()
    # Fetch other necessary data, like image
    image = db.session.query(Image).filter_by(id=requested_post.id).first()
    if image:
        image_data = image.data
    else:
        image_data = None

    return render_template("post.html", post=requested_post, image=image_data, logged_in=current_user.is_authenticated)


@app.route("/delete/<int:post_id>")
@login_required
def delete_post(post_id):
    requested_post = Post.query.filter_by(id=post_id).first()
    db.session.delete(requested_post)
    db.session.commit()
    flash("Deleted!")
    all_posts = Post.query.order_by(Post.date.desc()).all()
    return redirect(url_for('post', logged_in=current_user.is_authenticated))


@app.route('/')
def home():
    all_posts = Post.query.order_by(Post.date.desc()).all()
    todo_list = get_todo_list()
    # all_images = Post.query.order_by(Image.id.desc()).all()

    return render_template("index.html", todo_list=todo_list, all_posts=all_posts,
                           logged_in=current_user.is_authenticated)


@app.route('/add_todo', methods=["POST"])
def add_todo():
    todo_data = request.json
    todos.add_item(todo_data)
    all_posts = Post.query.order_by(Post.date.desc()).all()
    todo_list = get_todo_list()
    return render_template("added_todo.html", todo_list=todo_list, logged_in=current_user.is_authenticated)


@app.route('/delete_todo/<string:todo_id>', methods=["GET"])
def delete_todo(todo_id):
    todos.remove_item(todo_id)
    all_posts = Post.query.order_by(Post.date.desc()).all()
    todo_list = get_todo_list()
    return redirect(url_for('home', logged_in=current_user.is_authenticated))


@app.route('/logout', methods=["GET"])
def logout():
    logout_user()
    return redirect(url_for('home'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    # check_password_hash(pwhash, password)
    if request.method == "POST":
        user_email = request.form.get("email")
        user_password = request.form.get("password")

        # Querying the user based on the provided email
        user_check = User.query.filter_by(email=user_email).first()

        if user_check:
            if check_password_hash(user_check.password, user_password):
                login_user(user_check)
                return redirect(url_for("home", user_id=user_check.id, logged_in=current_user.is_authenticated))
            else:
                flash("Incorrect password, please try again.")
                return render_template("login.html", logged_in=current_user.is_authenticated)
        else:
            flash("That email does not exist, please try again.")
            return render_template("login.html", logged_in=current_user.is_authenticated)
    return render_template("login.html", logged_in=current_user.is_authenticated)


# @app.route("/ncaa", methods=["GET", "POST"])
# def ncaa():
#     form = XlForm()
#     picks = []
#     games = []
#     if form.validate_on_submit():
#         xl_file = form.excel_doc.data
#
#         try:
#             # Open the Excel file
#             wb = openpyxl.load_workbook(xl_file)
#
#             # Select the active worksheet
#             sheet = wb.active
#
#             # Grab items from columns D (column 4) and H (column 8) starting at the specified rows
#             column_data_D = [sheet.cell(row=i, column=4).value for i in range(9, 20)]
#             column_data_H = [sheet.cell(row=i, column=8).value for i in range(9, 20)]
#             column_data = zip(column_data_D, column_data_H)
#             # Remove numbering and store in a list
#             cleaned_column_data = []
#             for row in column_data:
#                 cleaned_row = []
#                 for item in row:
#                     if isinstance(item, str) and '. ' in item:
#                         cleaned_item = item.split('. ')[1]
#                     else:
#                         cleaned_item = item
#                     cleaned_row.append(cleaned_item)
#                 cleaned_column_data.append(cleaned_row)
#
#             picks = cleaned_column_data
#
#             import re
#
#             # Flatten the list of lists into a single list of strings
#             flattened_picks = [str(item) for sublist in picks for item in sublist]
#
#             # Replacements i.e. "Uconn" with "Connecticut"
#             flattened_picks = ["UConn" if pick == "Uconn" else pick for pick in flattened_picks]
#             flattened_picks = ["Colorado St." if pick == "Virginia/Colorado St." else pick for pick in flattened_picks]
#             flattened_picks = ["Saint Mary's (CA)" if pick == "Saint Mary's" else pick for pick in flattened_picks]
#             flattened_picks = ["Fla. Atlantic" if pick == "FAU" else pick for pick in flattened_picks]
#
#             # Replace "State" with "St."
#             flattened_picks = [re.sub(r'\bState\b', 'St.', pick) for pick in flattened_picks]
#             flattened_picks = ["NC State" if pick == "NC St." else pick for pick in flattened_picks]
#             print(f"picks pre-strip function={flattened_picks}")
#             flattened_picks = [pick.strip() for pick in flattened_picks]
#             print(f"after stripping: {flattened_picks}")
#
#             # Join picks into a comma-separated string
#             picks_str = ",".join(flattened_picks)
#
#             return redirect(url_for("ncaa_scoreboard", picks=picks_str))
#
#         except Exception as e:
#             flash("Error.")
#             print(f"Exception: {e}")
#             return render_template("ncaa.html", form=form)
#
#     return render_template("ncaa.html", form=form)


# @app.route("/ncaa-scoreboard/", methods=["GET", "POST"])
# def ncaa_scoreboard():
#     my_picks_str = request.args.get('picks')
#     my_picks = my_picks_str.split(',') if my_picks_str else []
#     games = get_games()
#     finished_games = []
#     correct = 0
#     incorrect = 0
#     all_done = False
#
#     # Get the current time in UTC
#     utc_now = datetime.utcnow()
#
#     # Define the time zone for Central Time
#     central_tz = pytz.timezone('America/Chicago')
#
#     # Convert UTC time to Central Time
#     central_now = utc_now.replace(tzinfo=pytz.utc).astimezone(central_tz)
#
#     # Extract only the date
#     central_today = central_now.date()
#     current_date = central_today.strftime('%A, %b %d')
#     games = [game for game in games if game['game']['bracketRound'] != ""]
#     for game in games:
#         if game['game']['gameState'] == "final":
#             pass
#             if int(game['game']['home']['score']) > int(game['game']['away']['score']):
#                 if game['game']['home']['names']['short'] in my_picks:
#                     correct += 1
#                 else:
#                     incorrect += 1
#             if int(game['game']['away']['score']) > int(game['game']['home']['score']):
#                 if game['game']['away']['names']['short'] in my_picks:
#                     correct += 1
#                 else:
#                     incorrect += 1
#
#         if game['game']['home']['score'] == "":
#             game['game']['home']['score'] = 0
#         else:
#             game['game']['home']['score'] = int(game['game']['home']['score'])
#
#         if game['game']['away']['score'] == "":
#             game['game']['away']['score'] = 0
#         else:
#             game['game']['away']['score'] = int(game['game']['away']['score'])
#
#         finished_games = [game for game in games if game['game']['gameState'] == "final"]
#         if len(finished_games) == len(games):
#             all_done = True
#
#     return render_template("ncaa_scoreboard.html", picks=my_picks, games=games,
#                            finished_games=finished_games, correct=correct, incorrect=incorrect, all_done=all_done,
#                            current_date=current_date)


def load_todo_list():
    """Load the to-do list from the JSON file."""
    try:
        with open(TODO_FILE, 'r') as file:
            todo_list = json.load(file)
    except FileNotFoundError:
        # If the file doesn't exist yet, return an empty list
        todo_list = []
    return todo_list


def save_todo_list(todo_list):
    """Save the to-do list to the JSON file."""
    with open(TODO_FILE, 'w') as file:
        json.dump(todo_list, file, indent=4)


def add_todo_item(title, description):
    """Add a to-do item to the to-do list."""
    todo_list = load_todo_list()
    todo_list.append({'title': title, 'description': description})
    save_todo_list(todo_list)


def get_todo_list():
    """Get the entire to-do list."""
    return load_todo_list()


@app.route('/work', methods=['GET'])
def work():

    return render_template("work.html", logged_in=current_user.is_authenticated)


if __name__ == "__main__":
    app.run(debug=True)
